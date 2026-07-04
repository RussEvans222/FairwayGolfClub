from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ===== SUMMARY SHEET =====
ws = wb.active
ws.title = "Budget Summary"

# Styling
header_font = Font(name='Arial', bold=True, size=14, color='1E2A3A')
sub_font = Font(name='Arial', bold=True, size=11, color='1E2A3A')
gold_font = Font(name='Arial', bold=True, size=12, color='B8860B')
blue_font = Font(name='Arial', size=10, color='0000FF')  # inputs
black_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
small_font = Font(name='Arial', size=9, color='666666')
header_fill = PatternFill('solid', fgColor='1E2A3A')
header_text = Font(name='Arial', bold=True, size=10, color='FFFFFF')
gold_fill = PatternFill('solid', fgColor='F5F0E0')
light_fill = PatternFill('solid', fgColor='F8F9FA')
thin_border = Border(
    bottom=Side(style='thin', color='CCCCCC')
)
thick_border = Border(
    top=Side(style='medium', color='1E2A3A'),
    bottom=Side(style='medium', color='1E2A3A')
)

# Column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 32

# Title
ws['B1'] = 'FAIRWAY GOLF CLUB'
ws['B1'].font = Font(name='Arial', bold=True, size=18, color='B8860B')
ws['B2'] = 'Phase 1 Budget — Value-Engineered Build'
ws['B2'].font = Font(name='Arial', size=12, color='666666')
ws['B3'] = 'Target: Under $200K | 2 Bays | Premium Aesthetic'
ws['B3'].font = Font(name='Arial', italic=True, size=10, color='1E2A3A')

# Headers
row = 5
for col, val in enumerate(['', 'Category', 'Budget', 'Full Premium', 'Smart Swap / Notes'], 1):
    c = ws.cell(row=row, column=col, value=val)
    if col >= 2:
        c.font = header_text
        c.fill = header_fill
        c.alignment = Alignment(horizontal='left' if col in [2,5] else 'right')

# ===== DATA =====
categories = [
    # (category, budget, premium, notes)
    ('SIMULATOR EQUIPMENT', None, None, None),
    ('Uneekor EYE XO2 launch monitors (x2)', 22000, 22000, 'Non-negotiable — best API + accuracy'),
    ('Projectors — Optoma GT2160HDR short-throw (x2)', 4000, 7000, '$2K ea vs $3.5K BenQ cinema-grade'),
    ('Impact screens — Carl\'s Place premium (x2)', 1600, 3000, '$800 ea; performs like $1,500 screens'),
    ('Enclosure frames — DIY steel tube (x2)', 1200, 4000, 'Steel tube + local welding vs pre-fab'),
    ('Gaming PCs — custom build (x2)', 4000, 6000, 'Build vs buy; i7 + RTX 4070 is plenty'),
    ('Software licenses (GSPro + Uneekor)', 1500, 1500, 'GSPro at $250/yr; Uneekor View included'),
    ('Hitting mats — Fiberbuilt (x2)', 2400, 2400, 'Don\'t cheap out — joint protection'),
    ('Auto-tee system — ProTee (x2)', 6000, 6000, 'Key differentiator; includes ball return'),
    ('Side monitors — data display (x2)', 800, 1600, '32" 4K from Costco vs commercial displays'),
    ('', None, None, None),
    ('BAY CONSTRUCTION', None, None, None),
    ('Bay framing + drywall (2 bays)', 8000, 12000, 'Standard stud framing; 12x15 each'),
    ('Acoustic panels — DIY mineral wool + fabric', 2400, 8000, '$300/bay DIY vs $4K/bay commercial'),
    ('Turf flooring — SynLawn putting grade', 3200, 5000, 'Direct-buy roll vs installed premium'),
    ('Sloped ball return floor — custom build', 3000, 5000, 'Concrete slope + turf overlay'),
    ('Electrical — dedicated circuits + lighting', 5000, 8000, 'Licensed electrician, just the bays'),
    ('HVAC — mini-split per bay', 4000, 6000, 'One 12K BTU unit serves both bays'),
    ('Bay divider wall', 1500, 3000, 'Framed wall, acoustic insulation'),
    ('LED strip lighting — bay perimeter', 600, 2000, 'Amazon LED strips vs custom install'),
    ('', None, None, None),
    ('LOUNGE & RECEPTION', None, None, None),
    ('Chesterfield sofa (x1 real + x1 quality repro)', 2800, 12000, 'FB Marketplace + Article.com vs RH'),
    ('Coffee table — reclaimed wood', 400, 2000, 'FB Marketplace / estate sale'),
    ('Reception desk — local woodworker live-edge', 2500, 8000, 'Instagram craftsman vs commercial'),
    ('Bar top — butcher block + dark stain + live-edge cut', 1200, 8000, 'Floor & Decor slab vs custom mill'),
    ('Bar stools (x4) — faux leather', 800, 3000, 'Amazon commercial-grade vs designer'),
    ('Walnut veneer slat wall — reception accent', 1800, 6000, 'Veneer panels vs solid walnut'),
    ('Bookshelf wall — IKEA hack + walnut stain', 400, 3000, 'KALLAX + stain + styled books'),
    ('Lighting — Edison pendants + gold ring fixture', 1200, 5000, 'Amazon/Wayfair vs designer fixtures'),
    ('Paint — navy ceiling, warm white walls', 800, 800, 'DIY weekend project'),
    ('Flooring — luxury vinyl plank (lounge area)', 3000, 8000, 'LVP wide-plank walnut-look vs hardwood'),
    ('Curtains + window treatments', 600, 2000, 'IKEA linen panels'),
    ('Plants + styling + decor', 800, 3000, 'Mix of real + faux; thrift frames'),
    ('', None, None, None),
    ('BEVERAGE SETUP', None, None, None),
    ('4-tap draft system — complete', 2500, 5000, 'Kegco or EdgeStar vs custom glycol'),
    ('Kegerator — dual-body commercial', 1500, 3000, 'Craigslist commercial or scratch-dent'),
    ('Glassware + bar supplies', 500, 1500, 'Restaurant supply store'),
    ('Mini fridge — BYOB holding', 300, 300, ''),
    ('ABC license application', 500, 500, 'VA retail off-premises; ~4-6 week process'),
    ('', None, None, None),
    ('TECHNOLOGY & SYSTEMS', None, None, None),
    ('Salesforce platform (initial setup)', 0, 0, 'Free tier to start; upgrade at 50+ members'),
    ('Booking system — Calendly/Square', 0, 500, 'Free tier handles 2 bays easily'),
    ('Digital displays — leaderboard + welcome (x2)', 1000, 4000, 'Consumer TVs + Raspberry Pi vs commercial'),
    ('Network + WiFi — Ubiquiti setup', 500, 1500, 'One AP + switch; overkill for 2 bays'),
    ('Security cameras (x3)', 400, 1500, 'Reolink PoE vs commercial NVR'),
    ('POS system — Square', 0, 2000, 'Free hardware with Square processing'),
    ('', None, None, None),
    ('SIGNAGE & BRANDING', None, None, None),
    ('Exterior sign — channel letters', 3500, 15000, 'Local sign shop vs architectural firm'),
    ('Logo wall — vinyl + LED backlight', 800, 5000, 'Vinyl on painted MDF + LED strip vs CNC metal'),
    ('Interior wayfinding + bay numbers', 300, 1500, 'Vinyl decals + brushed gold number plates'),
    ('', None, None, None),
    ('SOFT COSTS & WORKING CAPITAL', None, None, None),
    ('Permits + inspections', 2000, 3000, 'Fairfax County commercial build'),
    ('Insurance (6 months prepaid)', 3000, 3000, 'GL + property + liquor liability'),
    ('Legal — lease review + LLC', 2000, 2000, ''),
    ('Marketing — launch campaign', 2000, 5000, 'Instagram + Google Ads + local PR'),
    ('Working capital (3 months buffer)', 15000, 25000, 'Covers rent + utilities + payroll gap'),
    ('Contingency (10%)', None, None, 'Formula: 10% of total above'),
]

row = 6
section_rows = []
item_rows = []

for cat, budget, premium, notes in categories:
    row += 1
    if budget is None and premium is None and cat == '':
        continue  # skip spacer
    if budget is None and premium is None and cat != '':
        # Section header
        ws.cell(row=row, column=2, value=cat).font = gold_font
        section_rows.append(row)
    else:
        ws.cell(row=row, column=2, value=cat).font = black_font
        if budget is not None:
            c = ws.cell(row=row, column=3, value=budget)
            c.font = blue_font
            c.number_format = '$#,##0'
        if premium is not None:
            c = ws.cell(row=row, column=4, value=premium)
            c.font = small_font
            c.number_format = '$#,##0'
        if notes:
            ws.cell(row=row, column=5, value=notes).font = small_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=5).border = thin_border
        item_rows.append(row)

# Totals
row += 2
# Subtotal formula (sum all budget column entries)
budget_cells = ','.join([f'C{r}' for r in item_rows if ws.cell(row=r, column=3).value is not None and ws.cell(row=r, column=2).value != 'Contingency (10%)'])
premium_cells = ','.join([f'D{r}' for r in item_rows if ws.cell(row=r, column=4).value is not None and ws.cell(row=r, column=2).value != 'Contingency (10%)'])

subtotal_row = row
ws.cell(row=row, column=2, value='SUBTOTAL (before contingency)').font = bold_font
ws.cell(row=row, column=3, value=f'=SUM({budget_cells})').font = bold_font
ws.cell(row=row, column=3).number_format = '$#,##0'
ws.cell(row=row, column=4, value=f'=SUM({premium_cells})').font = small_font
ws.cell(row=row, column=4).number_format = '$#,##0'

row += 1
contingency_row = row
ws.cell(row=row, column=2, value='Contingency (10%)').font = black_font
ws.cell(row=row, column=3, value=f'=C{subtotal_row}*0.10').font = blue_font
ws.cell(row=row, column=3).number_format = '$#,##0'
ws.cell(row=row, column=4, value=f'=D{subtotal_row}*0.10').font = small_font
ws.cell(row=row, column=4).number_format = '$#,##0'

row += 1
total_row = row
ws.cell(row=row, column=2, value='TOTAL PHASE 1 INVESTMENT').font = Font(name='Arial', bold=True, size=12, color='1E2A3A')
ws.cell(row=row, column=3, value=f'=C{subtotal_row}+C{contingency_row}').font = Font(name='Arial', bold=True, size=12, color='B8860B')
ws.cell(row=row, column=3).number_format = '$#,##0'
ws.cell(row=row, column=4, value=f'=D{subtotal_row}+D{contingency_row}').font = Font(name='Arial', bold=True, size=11, color='666666')
ws.cell(row=row, column=4).number_format = '$#,##0'
for col in range(2, 6):
    ws.cell(row=row, column=col).border = thick_border

row += 1
ws.cell(row=row, column=2, value='SAVINGS vs Full Premium').font = bold_font
ws.cell(row=row, column=3, value=f'=D{total_row}-C{total_row}').font = Font(name='Arial', bold=True, size=11, color='228B22')
ws.cell(row=row, column=3).number_format = '$#,##0'

row += 2
ws.cell(row=row, column=2, value='Legend:').font = bold_font
row += 1
ws.cell(row=row, column=2, value='Blue values = hardcoded inputs you can adjust').font = blue_font
row += 1
ws.cell(row=row, column=2, value='Gray values = "full premium" comparisons from renders').font = small_font

wb.save('Fairway_Golf_Club_Phase1_Budget_200K.xlsx')
print("Done")
