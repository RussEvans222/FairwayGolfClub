from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
INPUT_FONT = Font(name='Arial', size=10, color='0000FF')
FORMULA_FONT = Font(name='Arial', size=10, color='000000')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
SECTION_FONT = Font(name='Arial', size=11, bold=True, color='0D2137')
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='0D2137')
SUBTITLE_FONT = Font(name='Arial', size=10, italic=True, color='666666')
GREEN_FONT = Font(name='Arial', size=10, bold=True, color='2A6041')
GOLD_FONT = Font(name='Arial', size=11, bold=True, color='8B6914')
RED_FONT = Font(name='Arial', size=10, color='CC0000')

NAVY_FILL = PatternFill('solid', fgColor='0D2137')
GREEN_FILL = PatternFill('solid', fgColor='EBF4EE')
GOLD_FILL = PatternFill('solid', fgColor='FDF6E8')

DOLLAR_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%'
NUM_FMT = '#,##0'

# ═══════════════════════════════════════════════════════════════════
# SHEET 1: ASSUMPTIONS — fixed row mapping
# ═══════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Assumptions"
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 35

ws['B2'] = "FAIRWAY GOLF CLUB — Model Assumptions"
ws['B2'].font = TITLE_FONT
ws['B3'] = "Blue values = inputs you can change to run scenarios"
ws['B3'].font = SUBTITLE_FONT

# Fixed row layout — every assumption gets a specific row we can reference
# Row 5: section header
# Rows 6+: data
data = {
    # (row, label, value, format, note)
    5: ("PERSONAL INCOME", None, None, None),
    6: ("Current OTE (W-2 salary + commission)", 227000, DOLLAR_FMT, "Russell's current comp"),
    7: ("Annual raise assumption", 0.03, PCT_FMT, "Conservative 3% annual"),
    8: ("Estimated effective tax rate", 0.38, PCT_FMT, "Federal + VA state + FICA on W-2"),
    9: ("", None, None, None),
    10: ("BUSINESS — PHASE 1 (2 BAYS)", None, None, None),
    11: ("Number of bays (Phase 1)", 2, NUM_FMT, "Starting configuration"),
    12: ("Operating hours per day", 12, NUM_FMT, "10am–10pm"),
    13: ("Days open per year", 360, NUM_FMT, "5 days closed/maintenance"),
    14: ("Hourly rate (blended avg)", 55, DOLLAR_FMT, "Mix: walk-in $65, member $40, off-peak $45"),
    15: ("Year 1 avg utilization", 0.35, PCT_FMT, "Ramp from 20% to 50% over 12 months"),
    16: ("Year 2 avg utilization", 0.50, PCT_FMT, "Steady state, mature 2-bay operation"),
    17: ("Year 3+ avg utilization (orig bays)", 0.55, PCT_FMT, "Slight growth from reputation"),
    18: ("", None, None, None),
    19: ("ANCILLARY REVENUE (as % of bay revenue)", None, None, None),
    20: ("Memberships / AI coaching premium", 0.12, PCT_FMT, "Recurring revenue above bay hours"),
    21: ("Beverage / bar revenue", 0.08, PCT_FMT, "ABC license + Bunny Man partnership"),
    22: ("Events / corporate / merch / leagues", 0.05, PCT_FMT, "Corporate nights, tournaments"),
    23: ("", None, None, None),
    24: ("MONTHLY OPERATING COSTS", None, None, None),
    25: ("Rent to Workhouse", 3500, DOLLAR_FMT, "Negotiate; could be $3K-$4.5K"),
    26: ("Part-time staff (2 employees)", 4800, DOLLAR_FMT, "2×25hrs/wk×$18/hr + 15% burden"),
    27: ("Utilities (electric, HVAC, internet)", 1200, DOLLAR_FMT, "Sims draw power; good HVAC needed"),
    28: ("Insurance (GL + property)", 500, DOLLAR_FMT, "Commercial policy"),
    29: ("Software / sim subscriptions", 600, DOLLAR_FMT, "Trackman/FullSwing subs, Salesforce"),
    30: ("Marketing / digital ads", 800, DOLLAR_FMT, "Google, social, local sponsorships"),
    31: ("Supplies / maintenance / misc", 500, DOLLAR_FMT, "Balls, tees, cleaning, bulbs"),
    32: ("ABC license (amortized monthly)", 200, DOLLAR_FMT, "~$2,400 one-time / 12 months"),
    33: ("", None, None, None),
    34: ("FINANCING", None, None, None),
    35: ("Total startup capital needed", 150000, DOLLAR_FMT, "Phase 1 value-engineered budget"),
    36: ("Owner equity (personal investment)", 75000, DOLLAR_FMT, "Cash from savings"),
    37: ("Loan amount (SBA or bank)", 75000, DOLLAR_FMT, "Financed portion"),
    38: ("Loan interest rate (annual)", 0.08, PCT_FMT, "SBA 7(a) typical rate"),
    39: ("Loan term (years)", 5, NUM_FMT, "60-month term"),
    40: ("", None, None, None),
    41: ("PHASE 2 EXPANSION (begins Year 3)", None, None, None),
    42: ("Additional bays", 2, NUM_FMT, "Expand from 2 → 4 total"),
    43: ("Phase 2 capital required", 70000, DOLLAR_FMT, "Sims + bay buildout only"),
    44: ("Phase 2 loan amount", 70000, DOLLAR_FMT, "Finance 100% from cash flow or loan"),
    45: ("Phase 2 new bays Year 1 utilization", 0.40, PCT_FMT, "Ramp faster (brand established)"),
    46: ("Phase 2 new bays steady state util", 0.55, PCT_FMT, "Match original bays"),
    47: ("Phase 2 incremental monthly costs", 1800, DOLLAR_FMT, "Extra power, sim subs, add'l staff hrs"),
}

for row, (label, value, fmt, note) in data.items():
    if label.isupper() and value is None:
        ws.cell(row, 2, label).font = SECTION_FONT
        ws.cell(row, 2).fill = GREEN_FILL
        for c in range(3, 5): ws.cell(row, c).fill = GREEN_FILL
    elif label:
        ws.cell(row, 2, label).font = FORMULA_FONT
        if value is not None:
            c = ws.cell(row, 3, value)
            c.font = INPUT_FONT
            if fmt: c.number_format = fmt
        if note:
            ws.cell(row, 4, note).font = SUBTITLE_FONT

# ═══════════════════════════════════════════════════════════════════
# SHEET 2: 5-YEAR PROJECTION
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("5-Year Projection")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 3
ws2.column_dimensions['B'].width = 42
for col in range(3, 8): ws2.column_dimensions[get_column_letter(col)].width = 16

ws2['B2'] = "5-YEAR HOUSEHOLD INCOME PROJECTION"
ws2['B2'].font = TITLE_FONT
ws2['B3'] = "Keep day job + Fairway as side business → evaluate full-time at Phase 2"
ws2['B3'].font = SUBTITLE_FONT

A = "Assumptions!"
# Monthly opex sum (rows 25-32)
OPEX_M = f"({A}C25+{A}C26+{A}C27+{A}C28+{A}C29+{A}C30+{A}C31+{A}C32)"
# PMT for Phase 1 loan: =PMT(rate/12, term*12, -amount)*12
PMT1 = f"PMT({A}C38/12,{A}C39*12,-{A}C37)*12"
# PMT for Phase 2 loan
PMT2 = f"PMT({A}C38/12,{A}C39*12,-{A}C44)*12"

# Layout rows
R = 5  # header row
# Rows relative to R:
# R+0 = 5: headers (Year 1..5)
# R+1 = 6: blank for section
rows = {}
rn = 6  # start data

def section(label):
    global rn
    rows[f'sec_{rn}'] = rn
    ws2.cell(rn, 2, label).font = SECTION_FONT
    ws2.cell(rn, 2).fill = GREEN_FILL
    for c in range(3, 8): ws2.cell(rn, c).fill = GREEN_FILL
    rn += 1

def row_label(key, label, font=FORMULA_FONT, fill=None):
    global rn
    rows[key] = rn
    ws2.cell(rn, 2, label).font = font
    if fill:
        for c in range(2, 8): ws2.cell(rn, c).fill = fill
    rn += 1

def blank():
    global rn
    rn += 1

# Build layout
section("YOUR W-2 INCOME")
row_label("w2", "Day job OTE (pre-tax)")
row_label("w2_post", "Day job after-tax take-home")
blank()
section("FAIRWAY GOLF CLUB")
row_label("bays", "Operating bays")
row_label("util", "Avg utilization (blended)")
row_label("bay_rev", "Bay rental revenue")
row_label("ancillary", "Ancillary revenue (members, bar, events)")
row_label("total_rev", "Total Fairway Revenue", GREEN_FONT)
blank()
row_label("opex", "Operating costs (annual)")
row_label("debt", "Loan payments (annual P&I)")
row_label("total_cost", "Total Fairway Costs", RED_FONT)
blank()
row_label("net", "FAIRWAY NET PROFIT (pre-tax)", GREEN_FONT, GOLD_FILL)
row_label("net_monthly", "  → Monthly average", FORMULA_FONT)
blank()
section("COMBINED HOUSEHOLD")
row_label("combined_pre", "Day Job + Fairway (pre-tax total)", GOLD_FONT, GOLD_FILL)
row_label("surplus", "Surplus vs. day job alone", GREEN_FONT)
blank()
row_label("quit_income", "IF RUSSELL QUITS: Fairway-only income")
row_label("quit_gap", "  → Gap vs. current $227K OTE")
row_label("quit_ready", "  → Ready to quit day job?", GREEN_FONT)
blank()
section("WEALTH BUILDING")
row_label("cum_profit", "Cumulative Fairway net profit")
row_label("equity_back", "Personal equity ($75K) recovered?")
row_label("biz_value", "Est. business value (3× annual profit)")
row_label("total_wealth", "Total new wealth created (cumulative + value)")

# Headers
for i, h in enumerate(["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]):
    c = ws2.cell(R, 2+i, h)
    c.font = HEADER_FONT; c.fill = NAVY_FILL; c.alignment = Alignment(horizontal='center')

# Fill formulas
for yi in range(5):
    col = 3 + yi
    cl = get_column_letter(col)
    yr = yi + 1

    # W-2 pre-tax
    ws2.cell(rows["w2"], col, f'={A}C6*(1+{A}C7)^{yi}')
    ws2.cell(rows["w2"], col).font = INPUT_FONT
    ws2.cell(rows["w2"], col).number_format = DOLLAR_FMT

    # W-2 after-tax
    ws2.cell(rows["w2_post"], col, f'={cl}{rows["w2"]}*(1-{A}C8)')
    ws2.cell(rows["w2_post"], col).number_format = DOLLAR_FMT

    # Bays
    if yr <= 2:
        ws2.cell(rows["bays"], col, f'={A}C11')
    else:
        ws2.cell(rows["bays"], col, f'={A}C11+{A}C42')
    ws2.cell(rows["bays"], col).number_format = NUM_FMT

    # Utilization
    if yr == 1:
        ws2.cell(rows["util"], col, f'={A}C15')
    elif yr == 2:
        ws2.cell(rows["util"], col, f'={A}C16')
    elif yr == 3:
        # Blend: orig bays at Y3 util + new bays at Phase 2 Y1 util
        ws2.cell(rows["util"], col, f'=({A}C11*{A}C17+{A}C42*{A}C45)/({A}C11+{A}C42)')
    elif yr == 4:
        ws2.cell(rows["util"], col, f'=({A}C11*{A}C17+{A}C42*{A}C46)/({A}C11+{A}C42)')
    else:
        ws2.cell(rows["util"], col, f'={A}C46')
    ws2.cell(rows["util"], col).number_format = PCT_FMT

    # Bay revenue = bays × hours × days × rate × util
    ws2.cell(rows["bay_rev"], col, f'={cl}{rows["bays"]}*{A}C12*{A}C13*{A}C14*{cl}{rows["util"]}')
    ws2.cell(rows["bay_rev"], col).number_format = DOLLAR_FMT

    # Ancillary
    ws2.cell(rows["ancillary"], col, f'={cl}{rows["bay_rev"]}*({A}C20+{A}C21+{A}C22)')
    ws2.cell(rows["ancillary"], col).number_format = DOLLAR_FMT

    # Total rev
    ws2.cell(rows["total_rev"], col, f'={cl}{rows["bay_rev"]}+{cl}{rows["ancillary"]}')
    ws2.cell(rows["total_rev"], col).font = GREEN_FONT
    ws2.cell(rows["total_rev"], col).number_format = DOLLAR_FMT

    # Opex
    if yr <= 2:
        ws2.cell(rows["opex"], col, f'={OPEX_M}*12')
    else:
        ws2.cell(rows["opex"], col, f'=({OPEX_M}+{A}C47)*12')
    ws2.cell(rows["opex"], col).number_format = DOLLAR_FMT

    # Debt
    if yr <= 2:
        ws2.cell(rows["debt"], col, f'={PMT1}')
    else:
        ws2.cell(rows["debt"], col, f'={PMT1}+{PMT2}')
    ws2.cell(rows["debt"], col).number_format = DOLLAR_FMT

    # Total cost
    ws2.cell(rows["total_cost"], col, f'={cl}{rows["opex"]}+{cl}{rows["debt"]}')
    ws2.cell(rows["total_cost"], col).font = RED_FONT
    ws2.cell(rows["total_cost"], col).number_format = DOLLAR_FMT

    # Net profit
    ws2.cell(rows["net"], col, f'={cl}{rows["total_rev"]}-{cl}{rows["total_cost"]}')
    ws2.cell(rows["net"], col).font = GREEN_FONT
    ws2.cell(rows["net"], col).fill = GOLD_FILL
    ws2.cell(rows["net"], col).number_format = DOLLAR_FMT

    # Monthly avg
    ws2.cell(rows["net_monthly"], col, f'={cl}{rows["net"]}/12')
    ws2.cell(rows["net_monthly"], col).number_format = DOLLAR_FMT

    # Combined pre-tax
    ws2.cell(rows["combined_pre"], col, f'={cl}{rows["w2"]}+{cl}{rows["net"]}')
    ws2.cell(rows["combined_pre"], col).font = GOLD_FONT
    ws2.cell(rows["combined_pre"], col).fill = GOLD_FILL
    ws2.cell(rows["combined_pre"], col).number_format = DOLLAR_FMT

    # Surplus
    ws2.cell(rows["surplus"], col, f'={cl}{rows["net"]}')
    ws2.cell(rows["surplus"], col).font = GREEN_FONT
    ws2.cell(rows["surplus"], col).number_format = DOLLAR_FMT

    # Quit income
    ws2.cell(rows["quit_income"], col, f'={cl}{rows["net"]}')
    ws2.cell(rows["quit_income"], col).number_format = DOLLAR_FMT

    # Quit gap
    ws2.cell(rows["quit_gap"], col, f'={cl}{rows["quit_income"]}-{A}C6')
    ws2.cell(rows["quit_gap"], col).number_format = DOLLAR_FMT

    # Quit ready
    ws2.cell(rows["quit_ready"], col, f'=IF({cl}{rows["quit_income"]}>={A}C6,"YES - Fairway alone covers OTE","Not yet - keep day job")')
    ws2.cell(rows["quit_ready"], col).font = GREEN_FONT

    # Cumulative profit
    if yi == 0:
        ws2.cell(rows["cum_profit"], col, f'={cl}{rows["net"]}')
    else:
        prev = get_column_letter(col-1)
        ws2.cell(rows["cum_profit"], col, f'={prev}{rows["cum_profit"]}+{cl}{rows["net"]}')
    ws2.cell(rows["cum_profit"], col).number_format = DOLLAR_FMT

    # Equity recovered
    ws2.cell(rows["equity_back"], col, f'=IF({cl}{rows["cum_profit"]}>={A}C36,"YES - equity recovered","Not yet")')
    ws2.cell(rows["equity_back"], col).font = GREEN_FONT

    # Biz value (3× net)
    ws2.cell(rows["biz_value"], col, f'=MAX(0,{cl}{rows["net"]}*3)')
    ws2.cell(rows["biz_value"], col).number_format = DOLLAR_FMT

    # Total wealth
    ws2.cell(rows["total_wealth"], col, f'={cl}{rows["cum_profit"]}+{cl}{rows["biz_value"]}')
    ws2.cell(rows["total_wealth"], col).number_format = DOLLAR_FMT

# ═══════════════════════════════════════════════════════════════════
# SHEET 3: YEAR 1 MONTHLY
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Year 1 Monthly")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 3
ws3.column_dimensions['B'].width = 34
for col in range(3, 16): ws3.column_dimensions[get_column_letter(col)].width = 11

ws3['B2'] = "YEAR 1 MONTHLY — The Realistic Ramp"
ws3['B2'].font = TITLE_FONT
ws3['B3'] = "Day job continues. Fairway ramps from soft-open to steady state."
ws3['B3'].font = SUBTITLE_FONT

# Headers
R = 5
months = ["Mo 1","Mo 2","Mo 3","Mo 4","Mo 5","Mo 6","Mo 7","Mo 8","Mo 9","Mo 10","Mo 11","Mo 12","TOTAL"]
ws3.cell(R, 2, "").font = HEADER_FONT; ws3.cell(R, 2).fill = NAVY_FILL
for i, m in enumerate(months):
    c = ws3.cell(R, 3+i, m); c.font = HEADER_FONT; c.fill = NAVY_FILL; c.alignment = Alignment(horizontal='center')

util_ramp = [0.20, 0.25, 0.28, 0.32, 0.35, 0.38, 0.40, 0.43, 0.45, 0.48, 0.50, 0.52]

# Row layout
m_rows = {}
rn3 = 6

def s3_section(label):
    global rn3
    ws3.cell(rn3, 2, label).font = SECTION_FONT
    ws3.cell(rn3, 2).fill = GREEN_FILL
    for c in range(3, 16): ws3.cell(rn3, c).fill = GREEN_FILL
    rn3 += 1

def s3_row(key, label, font=FORMULA_FONT, fill=None):
    global rn3
    m_rows[key] = rn3
    ws3.cell(rn3, 2, label).font = font
    if fill:
        for c in range(2, 16): ws3.cell(rn3, c).fill = fill
    rn3 += 1

def s3_blank():
    global rn3; rn3 += 1

s3_section("INCOME (Day Job)")
s3_row("w2m", "W-2 monthly (pre-tax)")
s3_row("w2_take", "W-2 after-tax take-home")
s3_blank()
s3_section("FAIRWAY REVENUE")
s3_row("util", "Bay utilization", INPUT_FONT)
s3_row("bay", "Bay rental revenue")
s3_row("anc", "Ancillary revenue")
s3_row("trev", "Total Fairway monthly revenue", GREEN_FONT)
s3_blank()
s3_section("FAIRWAY COSTS")
s3_row("opex", "Operating costs")
s3_row("loan", "Loan payment (monthly)")
s3_row("tcost", "Total Fairway costs", RED_FONT)
s3_blank()
s3_row("fnet", "FAIRWAY MONTHLY NET PROFIT", GREEN_FONT, GOLD_FILL)
s3_blank()
s3_section("HOUSEHOLD MONTHLY CASH")
s3_row("hh_job", "Day job take-home")
s3_row("hh_fairway", "Fairway net contribution")
s3_row("hh_total", "Total household monthly cash", GOLD_FONT, GOLD_FILL)
s3_blank()
s3_section("SAFETY NET")
s3_row("worst", "If Fairway makes $0 this month, you still have:")

for mi in range(12):
    col = 3 + mi
    cl = get_column_letter(col)

    ws3.cell(m_rows["w2m"], col, f'={A}C6/12')
    ws3.cell(m_rows["w2m"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["w2_take"], col, f'={cl}{m_rows["w2m"]}*(1-{A}C8)')
    ws3.cell(m_rows["w2_take"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["util"], col, util_ramp[mi])
    ws3.cell(m_rows["util"], col).font = INPUT_FONT
    ws3.cell(m_rows["util"], col).number_format = PCT_FMT

    # Bay rev: bays × hours × 30 × rate × util
    ws3.cell(m_rows["bay"], col, f'={A}C11*{A}C12*30*{A}C14*{cl}{m_rows["util"]}')
    ws3.cell(m_rows["bay"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["anc"], col, f'={cl}{m_rows["bay"]}*({A}C20+{A}C21+{A}C22)')
    ws3.cell(m_rows["anc"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["trev"], col, f'={cl}{m_rows["bay"]}+{cl}{m_rows["anc"]}')
    ws3.cell(m_rows["trev"], col).font = GREEN_FONT
    ws3.cell(m_rows["trev"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["opex"], col, f'={OPEX_M}')
    ws3.cell(m_rows["opex"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["loan"], col, f'=PMT({A}C38/12,{A}C39*12,-{A}C37)')
    ws3.cell(m_rows["loan"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["tcost"], col, f'={cl}{m_rows["opex"]}+{cl}{m_rows["loan"]}')
    ws3.cell(m_rows["tcost"], col).font = RED_FONT
    ws3.cell(m_rows["tcost"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["fnet"], col, f'={cl}{m_rows["trev"]}-{cl}{m_rows["tcost"]}')
    ws3.cell(m_rows["fnet"], col).font = GREEN_FONT
    ws3.cell(m_rows["fnet"], col).fill = GOLD_FILL
    ws3.cell(m_rows["fnet"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["hh_job"], col, f'={cl}{m_rows["w2_take"]}')
    ws3.cell(m_rows["hh_job"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["hh_fairway"], col, f'={cl}{m_rows["fnet"]}')
    ws3.cell(m_rows["hh_fairway"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["hh_total"], col, f'={cl}{m_rows["hh_job"]}+{cl}{m_rows["hh_fairway"]}')
    ws3.cell(m_rows["hh_total"], col).font = GOLD_FONT
    ws3.cell(m_rows["hh_total"], col).fill = GOLD_FILL
    ws3.cell(m_rows["hh_total"], col).number_format = DOLLAR_FMT

    ws3.cell(m_rows["worst"], col, f'={cl}{m_rows["w2_take"]}')
    ws3.cell(m_rows["worst"], col).font = GREEN_FONT
    ws3.cell(m_rows["worst"], col).number_format = DOLLAR_FMT

# TOTAL column
tc = 15
for key in ["w2m","w2_take","bay","anc","trev","opex","loan","tcost","fnet","hh_job","hh_fairway","hh_total","worst"]:
    ws3.cell(m_rows[key], tc, f'=SUM(C{m_rows[key]}:N{m_rows[key]})')
    ws3.cell(m_rows[key], tc).font = Font(name='Arial', size=10, bold=True)
    ws3.cell(m_rows[key], tc).number_format = DOLLAR_FMT

# ═══════════════════════════════════════════════════════════════════
# SHEET 4: THE PITCH
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("The Pitch")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 3
ws4.column_dimensions['B'].width = 52
ws4.column_dimensions['C'].width = 22

ws4['B2'] = "FAIRWAY GOLF CLUB — The Case for Our Family"
ws4['B2'].font = TITLE_FONT
ws4['B3'] = "Bottom line: We keep our income. This adds to it. Here's how."
ws4['B3'].font = Font(name='Arial', size=11, italic=True, color='2A6041')

pitch_items = [
    (5, "THE PLAN", None, SECTION_FONT, GREEN_FILL),
    (6, "Russell keeps his $227K/yr day job — no income change", None, FORMULA_FONT, None),
    (7, "2 part-time staff run Fairway during the day ($18/hr)", None, FORMULA_FONT, None),
    (8, "Russell closes up evenings (~9-10pm) — lives 1 mile away", None, FORMULA_FONT, None),
    (9, "Household income stays the same. Fairway profit is ADDITIVE.", None, GREEN_FONT, None),
    (10, "", None, None, None),
    (11, "WHAT WE RISK", None, SECTION_FONT, GOLD_FILL),
    (12, "Personal cash invested:", f"='Assumptions'!C36", INPUT_FONT, None),
    (13, "Monthly loan payment:", f"=PMT(Assumptions!C38/12,Assumptions!C39*12,-Assumptions!C37)", FORMULA_FONT, None),
    (14, "Russell's evenings for close-up (~1 hr/night)", None, FORMULA_FONT, None),
    (15, "If entire business fails, total loss = equity invested (like a car)", None, SUBTITLE_FONT, None),
    (16, "", None, None, None),
    (17, "WHAT WE GAIN EACH YEAR (added to household)", None, SECTION_FONT, GOLD_FILL),
    (18, "Year 1 — Fairway adds to household:", f"='5-Year Projection'!C{rows['net']}", FORMULA_FONT, None),
    (19, "Year 2 — Fairway adds:", f"='5-Year Projection'!D{rows['net']}", GREEN_FONT, None),
    (20, "Year 3 — Expand to 4 bays, adds:", f"='5-Year Projection'!E{rows['net']}", GREEN_FONT, None),
    (21, "Year 4 — Steady 4-bay operation, adds:", f"='5-Year Projection'!F{rows['net']}", GREEN_FONT, None),
    (22, "Year 5 — Mature operation, adds:", f"='5-Year Projection'!G{rows['net']}", GOLD_FONT, None),
    (23, "", None, None, None),
    (24, "YEAR 5 TOTAL HOUSEHOLD INCOME:", f"='5-Year Projection'!G{rows['combined_pre']}", GOLD_FONT, GOLD_FILL),
    (25, "vs. Day Job alone in Year 5:", f"='5-Year Projection'!G{rows['w2']}", FORMULA_FONT, None),
    (26, "", None, None, None),
    (27, "THE SAFETY NET", None, SECTION_FONT, GREEN_FILL),
    (28, "If Fairway earns $0, household still earns:", f"=Assumptions!C6", GREEN_FONT, None),
    (29, "Worst case total loss (personal equity):", f"=Assumptions!C36", RED_FONT, None),
    (30, "That's less than one new car. And we own the equipment as assets.", None, SUBTITLE_FONT, None),
    (31, "", None, None, None),
    (32, "THE UPSIDE (Year 5)", None, SECTION_FONT, GREEN_FILL),
    (33, "Business estimated value (sellable):", f"='5-Year Projection'!G{rows['biz_value']}", GOLD_FONT, None),
    (34, "Cumulative profits earned:", f"='5-Year Projection'!G{rows['cum_profit']}", GREEN_FONT, None),
    (35, "Total new household wealth created:", f"='5-Year Projection'!G{rows['total_wealth']}", GOLD_FONT, GOLD_FILL),
    (36, "", None, None, None),
    (37, "WHEN COULD RUSSELL QUIT?", None, SECTION_FONT, GOLD_FILL),
    (38, "Fairway alone covers $227K: Year 4-5 (4 bays at 55%)", None, GOLD_FONT, None),
    (39, "Recommended: keep day job through Year 2, evaluate at expansion", None, FORMULA_FONT, None),
    (40, "Best case: quit Year 3 if utilization consistently >55%", None, FORMULA_FONT, None),
    (41, "No pressure — the business works AS A SIDE INCOME indefinitely", None, GREEN_FONT, None),
]

for row, label, value, font, fill in pitch_items:
    ws4.cell(row, 2, label)
    if font: ws4.cell(row, 2).font = font
    if fill:
        ws4.cell(row, 2).fill = fill
        ws4.cell(row, 3).fill = fill
    if value:
        ws4.cell(row, 3, value)
        ws4.cell(row, 3).number_format = DOLLAR_FMT
        if font: ws4.cell(row, 3).font = font

out = 'Fairway_Income_Projection.xlsx'
wb.save(out)
print(f"Saved: {out}")
